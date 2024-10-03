import openpyxl
import pandas as pd


df = pd.read_excel('../networks/networks_list.xlsx')

category={'All':[], 'Social':[],'Informational':[],'Biological':[],'Economic':[],'Technological':[],'Transportation':[]}

datasets = df.values
for dataset in datasets:
    if dataset[0].isdigit():
        dataset[0] = 'Benchmark_{}'.format(dataset[0])
    category['All'].append(dataset[0])
    category[dataset[1]].append(dataset[0])

metrics = ['prec', 'auc_prec', 'auc_pr', 'auc_roc', 'auc_mroc', 'ndcg', 'mcc', 'h_measure']
lps = ['CN', 'RA', 'JA', 'PA', 'CH2', 'CN3', 'RA3', 'CH3', 'LRW', 'SRW', 'KA', 'MFI', 'SR', 'NMF', 'DW', 'N2V',
       'GCN', 'GAT', 'SAGE', 'VGNAE']


def calculate_average(file_paths):
    base_file = openpyxl.load_workbook('single/{}_discriminability.xlsx'.format(file_paths[0]))
    base_sheets = base_file.sheetnames

    new_file = openpyxl.Workbook()
    new_sheets = []

    for sheet_name in base_sheets:
        new_sheet = new_file.create_sheet(title=sheet_name)
        new_sheets.append(new_sheet)

    for file_path in file_paths:
        current_file = openpyxl.load_workbook('single/{}_discriminability.xlsx'.format(file_path))

        for i, sheet_name in enumerate(base_sheets):
            current_sheet = current_file[sheet_name]
            new_sheet = new_sheets[i]

            for row in range(1, current_sheet.max_row + 1):
                for column in range(1, current_sheet.max_column + 1):
                    cell_value = current_sheet.cell(row=row, column=column).value
                    if isinstance(cell_value, (int, float)):
                        new_cell = new_sheet.cell(row=row, column=column)
                        if new_cell.value is None:
                            new_cell.value = cell_value
                        else:
                            new_cell.value += cell_value

    for sheet_name in base_sheets:
        new_sheet = new_file[sheet_name]
        for row in range(1, new_sheet.max_row + 1):
            for column in range(1, new_sheet.max_column + 1):
                new_cell = new_sheet.cell(row=row, column=column)
                new_cell.value /= len(file_paths)

    default_sheet = new_file['Sheet']
    new_file.remove(default_sheet)
    return new_file


for key,datasets in category.items():
    if not datasets:
        continue
    result_file = calculate_average(datasets)
    result_file.save('category/{}_discriminability.xlsx'.format(key))
